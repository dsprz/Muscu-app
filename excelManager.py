#pip install openpyxl pandas
#pip install xlrd
#pip install xlsxwriter
import pandas as pd
import openpyxl
import os
from openpyxl.styles import Alignment, Font, PatternFill, colors
from openpyxl.utils import get_column_letter
import re
from datetime import date

excelPath = r"./Sport2.xlsx"

class ExcelManager :

    def __init__(self):
        if not os.path.exists(excelPath):
            self.createExcelFile()
        self.createHeaders()

    def createExcelFile(self):
        self.wbToCreate = openpyxl.Workbook()
        self.wbToCreate.save(excelPath)
    
    def writeInExcel(self, 
                     dataframe, 
                     exerciceNumber,
                     mode = "a"):
        dataframe2 = []
        for data in dataframe:
            dataframe2.append(data[1:])
        df2 = pd.DataFrame(dataframe2, columns=["Séries", "Répétitions", "Poids en kg"])
        #print(df2)
        writer = pd.ExcelWriter(excelPath, 
                                     mode=mode,
                                     engine="openpyxl",
                                     if_sheet_exists="overlay" if mode == "a" else None)
        
        with writer:
            #print(df2)
            maxRow = writer.sheets["Sheet"].max_row if mode == "a" else 1
            #print(maxRow)
            df2.to_excel(writer, 
                        sheet_name="Sheet",
                        startrow= maxRow, #if writer.sheets["Sheet"].max_row!=1 else 0,
                        header=None, 
                        startcol=2,
                        index=False,
                        )
        newMaxRow = writer.sheets["Sheet"].max_row
        self.createHeaders()
        self.createIndexes(dataframe, 
                           exerciceNumber, 
                           maxRow+1,
                           newMaxRow)
        self.formattingColumns()
        self.collapseRows(maxRow+1, 
                          newMaxRow, date.today().strftime("%d/%m/%Y"))

    def createHeaders(self):
        wb = openpyxl.load_workbook(excelPath)
        worksheet = wb["Sheet"]
        cellA1 = worksheet.cell(row=1, column=1)
        cellB1 = worksheet.cell(row=1, column=2)
        cellC1 = worksheet.cell(row=1, column=3)
        cellD1 = worksheet.cell(row=1, column=4)
        cellE1 = worksheet.cell(row=1, column=5)

        if cellA1.value != "Date":
            cellA1.value = "Date"
            cellA1.alignment = Alignment(horizontal="center",
                                        vertical="center",
                                        )
            cellA1.font = Font(bold=True,
                            size=12)
        if cellB1.value != "Exercices":
            cellB1.value = "Exercices"
            cellB1.alignment = Alignment(horizontal="center",
                                        vertical="center",
                                        )
            cellB1.font = Font(bold=True,
                            size=12)
            
        if cellC1.value != "Séries":
            cellC1.value = "Séries"
            cellC1.alignment = Alignment(horizontal="center",
                                        vertical="center",
                                        )
            cellC1.font = Font(bold=True,
                            size=12)
            
        if cellD1.value != "Répétitions":
            cellD1.value = "Répétitions"
            cellD1.alignment = Alignment(horizontal="center",
                                        vertical="center",
                                        )
            cellD1.font = Font(bold=True,
                            size=12)

        if cellE1.value != "Poids en kg":
            cellE1.value = "Poids en kg"
            cellE1.alignment = Alignment(horizontal="center",
                                        vertical="center",
                                        )
            cellE1.font = Font(bold=True,
                            size=12)
            
        wb.save(excelPath)
        wb.close()

    def createIndexes(self, 
                      dataframe, 
                      exerciceNumber, 
                      maxRow, 
                      newMaxRow):
        wb = openpyxl.load_workbook(excelPath)
        worksheet = wb["Sheet"]
        #df = pd.DataFrame(dataframe, columns=["Exercice", "Séries", "Répétitions", "Poids en kg"])
        
        for i in range(exerciceNumber):
            worksheet.merge_cells(f"B{maxRow+4*i}:B{maxRow+3+4*i}")
            cellB = worksheet.cell(row=maxRow+4*i, column=2)
            cellB.value = dataframe[4*i][0]
            cellB.alignment = Alignment(horizontal="center",
                                        vertical="center")
                        
        startDateRow = maxRow
        worksheet.merge_cells(f"A{startDateRow}:A{newMaxRow}") #changerA2 par maxRow ?
        cellDate = worksheet.cell(row=startDateRow, column=1)
        cellDate.value = date.today().strftime("%d/%m/%Y")+"\nBODYBUILDING"
        worksheet[f"A{startDateRow}"].alignment = Alignment(horizontal="center",
                                              text_rotation=90,
                                              vertical="center",
                                              wrap_text=True
                                              )
        worksheet[f"A{startDateRow}"].font = Font(name="Poppins",
                                    bold=True,
                                    size=14,
                                    )
        wb.save(excelPath)
        wb.close()


    def collapseRows(self, 
                     startCell, 
                     endCell, 
                     date):
        wb = openpyxl.load_workbook(excelPath)
        worksheet = wb["Sheet"]
        indicationCell = worksheet.cell(endCell + 1, column=1)
        indicationCell.value = date
        indicationCell.alignment = Alignment(horizontal="center",
                                             vertical="center")
        indicationCell.font = Font(size=15,
                                   bold=True,)
        worksheet.row_dimensions.group(startCell, endCell, hidden=False)
        worksheet.merge_cells(f"A{endCell + 1}:E{endCell + 1}")
        redFill = PatternFill(start_color='FFFF0000',
                        end_color='FFFF0000',
                        fill_type='solid')
        indicationCell.fill = redFill
        """for row in range(startCell, endCell + 1):
            worksheet.row_dimensions[row].hidden=True
            worksheet.row_dimensions[row].outline_level=1"""
        wb.save(excelPath)
        wb.close()

    def formattingColumns(self):
        """Formatte la taille d'une colonne pour qu'elle s'adapte au string le plus long"""
        sizeAdjustement = 5
        wb = openpyxl.load_workbook(excelPath)
        worksheet = wb["Sheet"]
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))+ sizeAdjustement)) 
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value
        wb.save(excelPath)
        wb.close()

        """firstLine = self.df.iloc[:2]
        print(firstLine)"""

    """def writeInExcel(self, row, data):
        startColumn = 2
        centerAlignFormat = self.wb.add_format({"align": "center_across"})
        for value in data:
            if type(value)!=int:
                self.ws.write(row, startColumn, value, centerAlignFormat)
            else:
                self.ws.write(row, startColumn, value)
            startColumn+=1"""
    
    def formatting(self, listToFormat):
        pass    

    def returnExcelDataframe(self):
        return pd.read_excel(excelPath, engine='openpyxl')
    
    def modifyExcel(self, dataframe, startRow):

        dataframe2 = []
        for data in dataframe:
            dataframe2.append(data[1:])
        df2 = pd.DataFrame(dataframe2, columns=["Séries", "Répétitions", "Poids en kg"])
        writer = pd.ExcelWriter(excelPath, 
                                     mode="a",
                                     engine="openpyxl",
                                     if_sheet_exists="overlay")
        
        with writer:
            #print(df2)
            #maxRow = writer.sheets["Sheet"].max_row
            #print(maxRow)
            df2.to_excel(writer, 
                        sheet_name="Sheet",
                        startrow= startRow, #if writer.sheets["Sheet"].max_row!=1 else 0,
                        header=None, 
                        startcol=1,
                        index=False,
                        )

    def testToExcel(self):
        df = pd.read_excel(excelPath, engine='openpyxl')
        #df = df.fillna(method="ffill")
        #print(self.getWorkoutSessionList())
            
    def getWholeDatabaseToList(self):
        """Lis le fichier Excel actuel et transforme tout le fichier en dataframe, 
        puis transforme la dataframe en plusieurs listes exploitables
        """
        df = pd.read_excel(excelPath, engine='openpyxl')

        self.dateList = [date for date in df[df.columns[0]].values.tolist() if str(date)!="nan"]
        self.exerciseList = [exercise for exercise in df[df.columns[1]].values.tolist() if str(exercise)!="nan"]

        self.seriesList = df[df.columns[2]].values.tolist()

        self.repetitionsList = [repetitions for repetitions in df[df.columns[3]].values.tolist() if str(repetitions)!="nan"]
        self.massList = [mass for mass in df[df.columns[4]].values.tolist() if str(mass)!="nan"]

        self.seperateSeries()
    
    def seperateSeries(self):
        self.exerciseSeriesList = []
        intermediateList = []
        a0 = 0

        #transformer les nan en SEPARE

        for index, serie in enumerate(self.seriesList):
            if str(serie) == "nan":
                self.seriesList[index] = "SEPARE"

        for serie in self.seriesList:
            if serie!="SEPARE":
                a = int(re.search(r'\d+', serie).group())
                if a > a0:
                    intermediateList.append(serie)
                    a0 = a
                else:
                    #intermediateList.insert(0, serie)
                    self.exerciseSeriesList.append(intermediateList)
                    intermediateList = []
                    intermediateList.append(serie)
                    a0 = 0
            else:
                intermediateList.append("SEPARE")
                self.exerciseSeriesList.append(intermediateList)
                intermediateList = []
                a0 = 0
                #self.exerciseSeriesList.append("SEPARE")"""

        #print(self.exerciseSeriesList)
        #self.exerciseSeriesList[0].pop(0)
        #self.exerciseSeriesList.pop(len(self.exerciseSeriesList)-1)
        #self.exerciseSeriesList[-1].insert(0, "Série 1")
        #print(self.exerciseSeriesList)
        #print("###################################")
        
        #self.exerciseSeriesList.append(["Serie 1", "Serie 2", "Serie 3", "Serie 4", "Serie 5"])
        #print(self.exerciseSeriesList)
        """print("###############################################################")
        for index, value in enumerate(self.exerciseSeriesList):
            #print(f"{index} : {value}")
            if value == "SEPARE":
                self.exerciseSeriesList[index-1].append(value)
                self.exerciseSeriesList.remove(value)
        print(self.exerciseSeriesList)"""
            
        """for serie in self.exerciseSeriesList:
            print(serie)
            if serie == "SEPARE":
                index = self.exerciseSeriesList.index(serie) - 1
                self.exerciseSeriesList[index].append(serie)
                self.exerciseSeriesList.remove(serie)
                print(self.exerciseSeriesList)
                print("##################################")"""

        #print(self.exerciseSeriesList)
        #self.exerciseSeriesList[-1].append("SEPARE")
        
    def getDateList(self):
        self.getWholeDatabaseToList()
        return self.dateList
    
    def getExerciseList(self):
        self.getWholeDatabaseToList()
        return self.exerciseList

    def getRepetitionsList(self):
        self.getWholeDatabaseToList()
        return self.repetitionsList

    def getMassList(self):
        self.getWholeDatabaseToList()
        return self.massList

    def getExerciseSeriesList(self):
        self.getWholeDatabaseToList()
        return self.exerciseSeriesList

    def getWorkoutSessionList(self):
        self.getWholeDatabaseToList()
        workoutSessionList = []
        returnList = []
        oldIndex = 0
        for index, exerciseSeries in enumerate(self.exerciseSeriesList):
            #print(f"{index} : {exerciseSeries}")
            separe = False
            intermediateList = []
            intermediateList.append(self.exerciseList[index])
            if "SEPARE" in exerciseSeries:
                separe = True
                exerciseSeries.pop()
            intermediateList.append(exerciseSeries)
            intermediateList.append(self.repetitionsList[oldIndex:len(exerciseSeries)*(index+1)])
            intermediateList.append(self.massList[oldIndex:len(exerciseSeries)*(index+1)])
            oldIndex = len(exerciseSeries)*(index+1)
            workoutSessionList.append(intermediateList)
            if separe:
                returnList.append(workoutSessionList)
                workoutSessionList = []

         #workoutSessionList.append(intermediateList)
        #print(workoutSessionList)
        #print(len(returnList))
        return returnList



#excelManager = ExcelManager()
